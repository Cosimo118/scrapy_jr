
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

#返回一个所有需要重新请求的formdata start的list
def get_excel_list(path,sheet_name,rowNum,wait_spi):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    empty_td = []
    for i in range(1,rowNum):
        num = ws.cell(row= i,column = column_index_from_string('AH'))
        print(num.value)
        empty_td.append(int(num.value))
    result = list(set(wait_spi).difference(set(empty_td)))
    return result

begin_td = list(range(1,227412,10))
temp_wait_spi = get_excel_list(r'E:\vscode_workspace\scrapy\tutorial\tutorial\real_test_01.xlsx',"Sheet1",217595,begin_td)
result =list(map(lambda x:x-1,temp_wait_spi))

with open("key.txt","w")as f:
    for item in result:
        f.write(str(item)+"\n")

with open("key.txt",'r')as f:
    list_key = f.readlines()

file = []
for item in list_key:
    file.append(int(item.strip()))
print(file)
print(len(file))