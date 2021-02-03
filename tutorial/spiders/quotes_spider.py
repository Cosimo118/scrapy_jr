import scrapy
import json
import random
import time
import requests

import xlwt
import xlrd
from xlutils.copy import copy

import sys
import os
# BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# sys.path.insert(0, BASE)
# # from utils import get_excel_list as util
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

filename = 'real_test.txt'

def get_excel_list(path,sheet_name,rowNum,wait_spi):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    empty_td = []
    for i in range(1,rowNum):
        num = ws.cell(row= i,column = column_index_from_string('AH'))
        print(num.value)
        empty_td.append(int(num.value))
    result = list(set(wait_spi).difference(set(empty_td)))
    return result

class QuotesSpider(scrapy.Spider):
    name = "quotes"
    def start_requests(self):
        #url = 'http://xkz.cbirc.gov.cn/jr/getLicence.do?useState=3'
        url = 'http://xkz.cbirc.gov.cn/jr/getLicence.do?useState=3&organNo=&fatherOrganNo=&province=&orgAddress=&organType=&branchType=&fullName=&address=&flowNo=&setDateStr=undefined&exitDateStr=undefined&jrOrganPreproty='
        headers = {'User-agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'}
        start_time = time.time()
        end_time = time.time()
        flag = 0
        url_proxy="http://api.xiaoxiangdaili.com/ip/get?appKey=672062632446283776&appSecret=zJi0ztph&cnt=&wt=text"
        resp = requests.get(url_proxy).text
        proxy = {"proxy":'http://'+str(resp)}
        with open("key.txt",'r')as f:
            list_key = f.readlines()
        key_file = []
        for item in list_key:
            key_file.append(int(item.strip()))
        wait_spi =list(map(lambda x:x-1,key_file))
        for item in wait_spi:
            if flag is 15:
                print(f"-------remain {len(wait_spi)} to get-----------")
                diff = 10.0-(end_time-start_time)
                sleep_time = diff if diff>0 else 0
                time.sleep(sleep_time)
                resp = requests.get(url_proxy).text
                proxy = {"proxy":'http://'+str(resp)}
                start_time = time.time()
                flag = 0
            start = str(item)
            form_data = {'start': start, 'limit': '10'}
            time.sleep(random.random()/10)
            yield scrapy.FormRequest(url,formdata=form_data,callback = self.parse,meta = proxy)
            flag = flag+1
            end_time = time.time()

    def parse(self, response):
        #改成存进txt吧...
        with open(filename,"a") as f:
            output = json.loads(response.body)
            if(len(output['datas']) is 10):
                for items in output['datas']:
                    value_list = list(items.values())
                    for v in value_list:
                        f.write(str(v)+'\t')
                    f.write('\n')


        ## 这里存入excel
        # workbook = copy(xlrd.open_workbook(os.path.join(r'E:\vscode_workspace\scrapy\tutorial\tutorial',filename)))
        # worksheet = workbook.get_sheet(0)
        # output = json.loads(response.body)
        # for items in output['datas']:
        #     colNum = items['colIndex']
        #     value_list = list(items.values())
        #     for k in range(0,len(value_list)):
        #         worksheet.write(colNum,k,label=str(value_list[k]))
        # workbook.save(filename)
